"""Tabular ingestion: parse CSV/XLSX row-based into invoices, normalize to schema, store per row."""

import csv
import io
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any

from sqlalchemy.orm import Session

from app.models.invoice import Invoice
from app.repositories import invoice_repo
from app.services.money_sanity import check_amounts_sanity, get_money_sanity_threshold
from app.services.quality_checks import key_fields_quality_ok

# Map normalized header (lower, stripped) -> schema key
COLUMN_ALIASES = {
    "invoice_number": ["invoice number", "invoice_number", "inv no", "invoice #", "invoice no", "inv number"],
    "vendor_name": ["vendor name", "vendor_name", "vendor", "seller", "from", "company"],
    "invoice_date": ["invoice date", "invoice_date", "date", "invoice date date"],
    "subtotal_amount": ["subtotal", "subtotal_amount", "subtotal amount"],
    "tax_amount": ["tax", "tax_amount", "tax amount", "vat"],
    "total_amount": ["total", "total_amount", "total amount", "amount", "grand total"],
    "currency": ["currency", "curr", "ccy"],
    "payment_status": ["payment status", "payment_status", "status", "payment"],
}

SCHEMA_KEYS = list(COLUMN_ALIASES.keys())
ALLOWED_CURRENCIES = {"USD", "EUR", "GBP", "INR"}


def _normalize_header(header: str) -> str:
    """Map header to schema key or return empty if unknown."""
    key = (header or "").strip().lower().replace("-", " ").replace("_", " ")
    for schema_key, aliases in COLUMN_ALIASES.items():
        if key in aliases or key == schema_key.replace("_", " "):
            return schema_key
    return ""


def _build_column_index(headers: list[str]) -> dict[str, int]:
    """Return schema_key -> column index (0-based)."""
    index: dict[str, int] = {}
    for i, h in enumerate(headers):
        schema_key = _normalize_header(str(h).strip())
        if schema_key and schema_key not in index:
            index[schema_key] = i
    return index


def _safe_numeric(val: Any) -> float | None:
    """Convert to float safely; None for empty or invalid."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, Decimal):
        return float(val)
    try:
        s = str(val).strip().replace(",", "")
        return float(s) if s else None
    except (TypeError, ValueError):
        return None


def _parse_date(val: Any) -> date | None:
    """Parse value to date (not datetime). Returns date or None."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(s[:19], fmt)
            return dt.date()
        except (ValueError, TypeError):
            continue
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.date()
    except (ValueError, TypeError):
        return None


def _clamp_currency(val: Any) -> str:
    if not val:
        return "USD"
    c = str(val).strip().upper()[:3]
    return c if c in ALLOWED_CURRENCIES else "USD"


def _row_to_normalized(row: list, col_index: dict[str, int], source_file: str) -> tuple[dict, str, list[str]]:
    """
    Convert a data row to normalized invoice dict, processing_status, and warnings.
    Returns (dict for Invoice, processing_status, warnings).
    """
    def get(k: str, default=None):
        i = col_index.get(k)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return v if v is not None and str(v).strip() != "" else default

    invoice_number = str(get("invoice_number") or "").strip() or "TABULAR"
    vendor_name = str(get("vendor_name") or "").strip() or "Unknown Vendor"
    raw_date = get("invoice_date")
    invoice_date = _parse_date(raw_date) or date.today()
    subtotal_amount = _safe_numeric(get("subtotal_amount"))
    tax_amount = _safe_numeric(get("tax_amount"))
    total_amount = _safe_numeric(get("total_amount"))
    if (total_amount is None or total_amount == 0) and ((subtotal_amount or 0) > 0 or (tax_amount or 0) > 0):
        total_amount = (subtotal_amount or 0) + (tax_amount or 0)
    # Row-level warnings
    warnings: list[str] = []
    vendor_defaulted = not vendor_name or vendor_name == "Unknown Vendor"
    invoice_number_defaulted = not invoice_number or invoice_number == "TABULAR"
    total_missing_or_zero = total_amount is None or total_amount <= 0
    if vendor_defaulted:
        warnings.append("Vendor name defaulted")
    if invoice_number_defaulted:
        warnings.append("Invoice number defaulted")
    if total_missing_or_zero:
        warnings.append("Total missing")
    # SUCCESS: invoice_number and vendor_name not default AND total_amount > 0, and quality checks pass.
    # PARTIAL: row saved but one of (invoice_number/vendor_name) missing/weak OR warnings exist OR quality fail.
    # FAILED: total_amount=0 AND key fields missing (no good invoice_number or vendor_name).
    if total_missing_or_zero and (vendor_defaulted or invoice_number_defaulted):
        status = "FAILED"
    elif not vendor_defaulted and not invoice_number_defaulted and not total_missing_or_zero:
        status = "SUCCESS"
    else:
        status = "PARTIAL"
    if status == "SUCCESS" and not key_fields_quality_ok(invoice_number, vendor_name):
        status = "PARTIAL"
        warnings.append("Low-quality vendor/invoice fields")
    # Money sanity: if any amount exceeds threshold, mark PARTIAL and add warning
    s_amt = (subtotal_amount if subtotal_amount is not None else 0) or 0
    t_amt = (tax_amount if tax_amount is not None else 0) or 0
    tot_amt = (total_amount if total_amount is not None else 0) or 0
    sanity_warnings = check_amounts_sanity(
        {"subtotal_amount": s_amt, "tax_amount": t_amt, "total_amount": tot_amt},
        get_money_sanity_threshold(),
    )
    warnings.extend(sanity_warnings)
    if sanity_warnings and status == "SUCCESS":
        status = "PARTIAL"
    currency = _clamp_currency(get("currency"))
    payment_status = str(get("payment_status") or "").strip() or "pending"

    return (
        {
            "invoice_number": invoice_number,
            "vendor_name": vendor_name,
            "invoice_date": invoice_date,
            "subtotal_amount": Decimal(str(s_amt)),
            "tax_amount": Decimal(str(t_amt)),
            "total_amount": Decimal(str(tot_amt)),
            "currency": currency,
            "payment_status": payment_status,
            "source_file": source_file,
        },
        status,
        warnings,
    )


def _invoice_to_response(inv: Invoice) -> dict:
    """Created invoice as JSON with source_file and processing_status."""
    return {
        "id": inv.id,
        "invoice_number": inv.invoice_number,
        "vendor_name": inv.vendor_name,
        "invoice_date": inv.invoice_date.isoformat() if inv.invoice_date else None,
        "subtotal_amount": float(inv.subtotal_amount),
        "tax_amount": float(inv.tax_amount),
        "total_amount": float(inv.total_amount),
        "currency": inv.currency,
        "payment_status": inv.payment_status,
        "processing_status": inv.processing_status,
        "source_file": inv.source_file,
        "uploaded_at": inv.uploaded_at.isoformat() if inv.uploaded_at else None,
    }


def _rows_from_csv(file_bytes: bytes) -> tuple[list[str], list[list]]:
    """Return (headers, list of row lists)."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    headers = [str(h).strip() for h in rows[0]]
    data_rows = rows[1:]
    return headers, data_rows


def _rows_from_xlsx(file_bytes: bytes) -> tuple[list[str], list[list]]:
    """Return (headers, list of row lists) from first sheet."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = [[c for c in row] for row in rows[1:]]
    return headers, data_rows


def _process_tabular_impl(file_bytes: bytes, filename: str, content_type: str, db: Session) -> list[dict]:
    """Parse CSV or XLSX into rows, store each as invoice. Returns list of created invoice dicts with source_file, processing_status, warnings."""
    fn = (filename or "").lower()
    ct = (content_type or "").lower()
    if ".csv" in fn or "csv" in ct:
        headers, data_rows = _rows_from_csv(file_bytes)
    elif ".xlsx" in fn or "spreadsheet" in ct or "excel" in ct or "xlsx" in ct or "vnd.openxmlformats" in ct:
        headers, data_rows = _rows_from_xlsx(file_bytes)
    else:
        raise ValueError("Only CSV or XLSX are supported for tabular ingestion")

    col_index = _build_column_index(headers)
    if not col_index:
        raise ValueError("No recognizable invoice columns in file")

    created: list[dict] = []
    for row in data_rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        normalized, status, row_warnings = _row_to_normalized(row, col_index, filename)
        invoice = Invoice(
            invoice_number=normalized["invoice_number"],
            vendor_name=normalized["vendor_name"],
            invoice_date=normalized["invoice_date"],
            subtotal_amount=normalized["subtotal_amount"],
            tax_amount=normalized["tax_amount"],
            total_amount=normalized["total_amount"],
            currency=normalized["currency"],
            payment_status=normalized["payment_status"],
            processing_status=status,
            source_file=normalized["source_file"],
        )
        saved = invoice_repo.create(db, invoice)
        created.append({**_invoice_to_response(saved), "warnings": row_warnings})
    return created


class TabularIngestionService:
    """Parse CSV/XLSX into invoices (row-based). Returns list of created invoice dicts with source_file and processing_status."""

    def process_tabular(self, file_bytes: bytes, filename: str, content_type: str, db: Session) -> list[dict]:
        """Parse CSV or XLSX into rows, normalize columns to schema, store each row as an invoice."""
        return _process_tabular_impl(file_bytes, filename, content_type, db)
